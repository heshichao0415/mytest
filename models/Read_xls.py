import os
from configpath import getpath
from xlrd import open_workbook
import openpyxl
import json
import xlwt

class Case:
    def __init__(self):
        self.case_id = None
        self.case_name = None
        self.method = None
        self.url = None
        self.data = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None


class readExcel():
    def get_xls(self, xls_name, sheet_name):
        cls = []
        cls1 = []          #实验列表
        self.xlsPath = os.path.join(getpath(), "Case_data", xls_name)
        file = open_workbook(self.xlsPath, 'r')                  #打开工作簿
        sheet = file.sheet_by_name(sheet_name)              #通过名字打开工作簿
        rows = sheet.nrows
        for i in range(rows):
            if sheet.row_values(i)[0] != 'case_id':
                cls.append(sheet.row_values(i))
        return cls

    def write_xls(self, xls_name, sheet_name):
        list = self.get_xls('case_xy_huawei.xlsx', 'Sheet5')
        self.xlsPath = os.path.join(getpath(), "Case_data", xls_name)
        wb = openpyxl.load_workbook(self.xlsPath)
        ws = wb[sheet_name]
        for i in list:
            ws.cell(row=int(i[0]), column=1).value = int(i[0])
            ws.cell(row=int(i[0]), column=2).value = i[1]
        wb.save(self.xlsPath)
        # ws['A1'] = 24
        # ws.cell(row=1, column=3).value = 123
        # wb.save(self.xlsPath)



# class readExcel():
#     def get_xls(self, xls_name, sheet_name):
#         cases = []
#         xlspath = os.path.join(getpath(), "Case_data", xls_name)
#         file = openpyxl.load_workbook(xlspath)
#         sheet = file[sheet_name]
#         rows = sheet.max_row
#         for i in range(2, rows+1):
#             case = Case()
#             case_id = sheet.cell(row=i, column=1).value
#             case_name = sheet.cell(row=i, column=2).value
#             case_method = sheet.cell(row=i, column=3).value
#             case_url = sheet.cell(row=i, column=4).value
#             case_datas = sheet.cell(row=i, column=5).value
#             case_expected = sheet.cell(row=i, column=6).value
#             case.id = case_id
#             case.name = case_name
#             case.method = case_method
#             case.url = case_url
#             case.datas = case_datas
#             case.expected = case_expected
#             cases.append(case)
#         file.close()
#         return cases
if __name__ == "__main__":
    readExcel().write_xls('case_xy_huawei.xlsx', 'Sheet6')